# -*- coding: utf-8 -*-
"""primeiro exemplo em Pandas.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1Yto0xL5Sk0uTe7_lBWhwxfoJefyGdSte
"""

import pandas as pd

combustiveis_df = pd.read_excel ("ca-2021-02.xlsx")

# Usa o print para ver o dataframe!
print(combustiveis_df)

display(combustiveis_df)

#Exibe as primeiras 5 linhas
display(combustiveis_df.head())

# Quero, na verdade, exibir as primeiras 15 linhas
display(combustiveis_df.head(15))

print(combustiveis_df.shape)

print(combustiveis_df.shape[0])

#Quais são as colunas e que tipo de dados cada um tem...

print(combustiveis_df.info())

print(combustiveis_df.describe())

display(combustiveis_df['Revenda'])

#add .head(10) para ver somente as primeiras 10 linhas

ca_df = combustiveis_df[['Revenda', 'Municipio','Produto','Valor de Venda']]
display(ca_df)

#Exibe a 4a. linha.

display(ca_df.loc[3])

#Exibir da 10a. linha até a 20a.linha

display(ca_df.loc[9:19])

# Criar um dataframe gas_df contendo 
# apenas as 4 colunas (Revenda, Municipio, Produto, Valor de Venda)
# somente com combustível

gas_df = ca_df.loc[ca_df['Produto'] == 'GASOLINA']
display(gas_df)

display(gas_df['Valor de Venda'].max())

display(gas_df[['Revenda','Municipio','Valor de Venda']].max())

#DataFrame.loc[] com múltiplas condições para filtragem
#Quais são os preços, postos que vendem ETANOL na minha cidade (INDAIATUBA) 
#ordenado do menor valor de venda para o maior
etanol_indaiatuba_df = ca_df.loc[(ca_df['Produto'] == 'ETANOL') & (ca_df['Municipio'] == 'INDAIATUBA')]
etanol_indaiatuba_df.sort_values(by='Valor de Venda')
display(etanol_indaiatuba_df)

# Qual a média de preços dos combustíveis GASOLINA e GASOLINA ADITIVADA do Bairro MOOCA em SÃO PAULO?
display(combustiveis_df.loc[(combustiveis_df['Bairro'] == 'MOOCA') & 
                            (combustiveis_df['Municipio'] == 'SAO PAULO') & 
                            ((combustiveis_df['Produto'] == 'GASOLINA') | (combustiveis_df['Produto'] == 'GASOLINA ADITIVADA')), 
                            ['Valor de Venda']].mean())

# Como mostrar média de valor de venda POR COMBUSTÍVEL Brasil todo?
media_por_combustivel_df = ca_df[['Produto', 'Valor de Venda']].groupby(by='Produto').mean().round(2)
display(media_por_combustivel_df)

#Quero adicionar uma coluna de valor booleano no combustiveis_df
# chamada "Ativo" que, por padrão, vai ser True para 
# TODAS as linhas
combustiveis_df['Ativo'] = True
print(combustiveis_df.info())
display(combustiveis_df.head())

# Exportar para Excel o dataframe com etanol em Indaiatuba....
etanol_indaiatuba_df.to_excel('etanol_indaiatuba.xlsx', sheet_name='Etanol em Indaiatuba')

combustiveis_df = pd.read_excel("ca-2021-02.xlsx")
display(combustiveis_df.head())

#Inserção simples de dado
combustiveis_df['Ativo'] = True

display(combustiveis_df.head())

"""# Criar uma coluna "Obs" que tenha nela escrito "MELHOR CIDADE" quando a coluna Municipio for igual a SAO PAULO
combustiveis_df['Obs'] = ["MELHOR CIDADE" if municipio == 'SAO PAULO' else None for municipio in combustiveis_df['Municipio']]
display(combustiveis_df.loc[combustiveis_df['Municipio'].isin(['SAO PAULO','INDAIATUBA', 'CAMPINAS', 'SALTO']), ['Municipio', 'Obs']])
"""

# (por Leandro Rodrigues)
# como preencher uma coluna 'Valor de Venda - Status' que verifica o seguinte:
# se o valor de venda for maior que 6,5 reais, ele fala que tá Caro..caso contrário, está barato
import numpy as np

combustiveis_df['Status do Valor de Venda'] = np.where(combustiveis_df['Valor de Venda'] > 6.5, 'Caro', 'Barato')
display(combustiveis_df[['Revenda', 'Valor de Venda', 'Status do Valor de Venda']])

# Calcular postos de gasolina por habitante temos na amostragem de 
# combustiveis nov/2021

num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv", sep=";")
display(num_habitantes_df)

# Calcular postos de gasolina por habitante temos na amostragem de 
# combustiveis nov/2021

num_habitantes_df = pd.read_csv("ibge_num_habitantes_estimado.csv", sep=";")
num_habitantes_df.rename(columns={"Estado":"Estado - Sigla"}, inplace=True)
display(num_habitantes_df)

# Faz um MERGE dos dois dataframes
colunas = ['Municipio', 'Estado - Sigla']
merge_df = combustiveis_df.merge(num_habitantes_df, how="inner", on=colunas)
display(merge_df)
print(merge_df.info())

#Destruir coluna completamente vazia (todas as linhas são nulas)
merge_df.dropna(axis='columns', inplace=True)
print(merge_df.info())

colunas=['Regiao - Sigla', 'Nome da Rua', 'Numero Rua', 
         'Bairro', 'Cep', 'Produto', 'Data da Coleta', 'Valor de Venda',
         'Unidade de Medida', 'Bandeira', 'Ativo', 'Status do Valor de Venda']
merge_df.drop(labels=colunas, axis=1, inplace=True)
print(merge_df.info())

# Remover a linhas duplicadas
merge_df.drop_duplicates(inplace=True)
display(merge_df.head(100))

#Agrupar e contar quantos postos tem na cidade..
postos_por_municipio_df = merge_df.groupby(by=['Estado - Sigla', 'Municipio', 'NumHabitantes2021']).count()
postos_por_municipio_df.drop('CNPJ da Revenda', axis=1, inplace=True)
postos_por_municipio_df.rename(columns={"Revenda": "Número de Postos"}, inplace=True)
display(postos_por_municipio_df)

Agrupar e contar quantos postos tem na cidade..
postos_por_municipio_df = merge_df.groupby(by=['Estado - Sigla', 'Municipio', 'NumHabitantes2021']).count()
postos_por_municipio_df.reset_index(inplace=True)
#display(postos_por_municipio_df.info())
postos_por_municipio_df.drop('CNPJ da Revenda', axis=1, inplace=True)
postos_por_municipio_df.rename(columns={"Revenda": "NumPostos"}, inplace=True)

postos_por_municipio_df['PostosPorHabitante'] = postos_por_municipio_df['NumPostos'] / postos_por_municipio_df['NumHabitantes2021']
display(postos_por_municipio_df.info())
display(postos_por_municipio_df)

#Vamos brincar de gráficos!!!

import matplotlib.pyplot as plt

plt.hist(combustiveis_df['Valor de Venda'])
#Vamos colocar um título no gráfico
plt.title("Preço dos combustíveis - Nov/2021")
#Rótulo horizontal e vertical
plt.xlabel("Preço (em reais)")
plt.ylabel("Quantidade de Coletas")

#Traça a linha vermelha tracejada com o preço médio
plt.axvline(combustiveis_df['Valor de Venda'].mean(), color='red', linestyle='dashed', linewidth=5)


#"Plota" o gráfico
plt.show()

import seaborn as sns

#Vou definir a área do gráfico
plt.figure(figsize=(7,5))

#Plotar o gráfico
c_mean.plot(
    kind="barh",
    xlabel="Tipo de Combustível",
    ylabel="Preço reais/litro",
    title="Média de preços por combustível",
    color="red",
    alpha=0.3
)

#Grid
plt.grid()

#Remover as linhas superior e lateral direita do gráfico 
sns.despine()

#Exibe
plt.show()

#c_mean.reset_index(inplace=True)
display(c_mean)

excel = "por_litro.xlsx"
c_mean.to_excel(excel, "Sumário")

#Vamor personalizar a planilha Excel...
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Color, Alignment

#Vai abrir o Excel no openpyxl
wb = load_workbook(excel) # wb = Workbook

#Pegar a planilha certa... usando o Sheet Name (nome da planilha)
ws = wb['Sumário'] # Work Sheet -> planilha atual, ativa, de trabalho

#Vamos pintar o cabeçalho da tabela de "cinzinha"
cinzinha = PatternFill("solid", fgColor="CCCCCC")
coords = ['A1', 'B1']
for coord in coords:
  ws[coord].fill = cinzinha

#Onde o preço do combustível for maior ou igual a 6,5 reais (6.5) pinta a fonte
#de vermelho e deixa negrito...
MAX_ROW = ws.max_row
num_linha = 2
while (num_linha <= MAX_ROW):
  coord = 'B'+str(num_linha) #coord="B{0}".format(num_linha)
  if ws[coord].value >= 6.5:
    ws[coord].font = Font(bold=True, color="FF0000")
  num_linha = num_linha + 1
#Salvar o Excel
wb.save(excel)